#Description: This a Update on my last project. That project is about creating a School Management System  that will read, write and delete data from excel File using a python library called as openpyxl. In This Article we will make a login system for  school management, Saving pass and user id's in second sheet of the excel sheet, will bind keys to make entering the into app easier, changing color coading for regular use etc.

#Code:
import tkinter as tk
from tkinter import *
import os 
from os import path 
from openpyxl import *
from PIL import Image, ImageTk
from tkinter import messagebox

app = tk.Tk()
app.title('School Management System')
app.geometry('500x350')
menu = Menu(app)
app.configure(menu=menu, bg="#eaeaea")

#----------Opening Our Excle File------------ (path.join(path.dirname(__file__), "data","student.xlsx"))

try:
  file = load_workbook(path.join(path.dirname(__file__), "data","student.xlsx"))
  sheet = file.active
  sheet1 = file['Sheet1']
  sheet2 = file['Sheet2']
except:
  messagebox.showerror("Error", "File not found")

#---variables to store the user input Data----
login_user = StringVar()
login_pass = StringVar()
register_user = StringVar()
register_pass = StringVar()
register_auth = StringVar()

ent1_var = StringVar()
ent2_var = StringVar()
ent3_var = StringVar()
ent4_var = StringVar()
ent5_var = StringVar()
ent6_var = StringVar()
ent7_var = StringVar()

#---Class to save first line of excle Sheets-----


class First_line():
  def page(self):
    sheet1.column_dimensions['A'].width = 25
    sheet1.column_dimensions['B'].width = 25
    sheet1.column_dimensions['C'].width = 25
    sheet1.column_dimensions['D'].width = 15
    sheet1.column_dimensions['E'].width = 20
    sheet1.column_dimensions['F'].width = 25
    sheet1.column_dimensions['G'].width = 50

    sheet1.cell(row=1, column=1).value = "Name"
    sheet1.cell(row=1, column=2).value = "Father's name"
    sheet1.cell(row=1, column=3).value = "Enrollment no."
    sheet1.cell(row=1, column=4).value = "Class"
    sheet1.cell(row=1, column=5).value = "Contact Number"
    sheet1.cell(row=1, column=6).value = "Email id"
    sheet1.cell(row=1, column=7).value = "Address"

    sheet2.column_dimensions['A'].width = 25
    sheet2.column_dimensions['B'].width = 25
    sheet2.cell(row=1, column=1).value = "Userame"
    sheet2.cell(row=1, column=2).value = "Password"

    file.save(path.join(path.dirname(__file__), "data","student.xlsx"))
  #-----------------------------------------------

#----------Class for Login System-----------------


class Login_sys:

  #---Creating Front end for the Sign in option---
  def sign_in_frontend(self):
    login_win = Toplevel()
    login_win.title('Sign In')
    login_win.geometry('500x350')
    login_win.configure(bg="#eaeaea")

    msg1 = tk.Label(login_win, text='User Name', )
    msg1.place(relx=0.2, rely=0.3, anchor=CENTER)

    msg2 = tk.Label(login_win, text=' Password ', )
    msg2.place(relx=0.2, rely=0.5, anchor=S)

    user_name = Entry(login_win, textvariable=login_user, )
    user_name.place(relx=0.6, rely=0.3, anchor=CENTER, width=300)

    user_pass = Entry(login_win, show="*",
                      textvariable=login_pass, )
    user_pass.place(relx=0.6, rely=0.5, anchor=S, width=300,)

    button = tk.Button(login_win, text='Sign In', width=20,
                       height=2, command=self.sign_in_backend,
                       activebackground="dark grey", activeforeground="red")
    button.place(relx=0.5, rely=0.7, anchor=CENTER)

  #---Creating Front end for the Sign Up option---
  def sign_up_frontend(self):
    register_win = Toplevel()
    register_win.title('Sign Up')
    register_win.geometry('500x350')
    register_win.configure(bg="#eaeaea")

    msg1 = tk.Label(register_win, text='User Name')
    msg1.place(relx=0.2, rely=0.2, anchor=CENTER)

    msg2 = tk.Label(register_win, text=' Password ')
    msg2.place(relx=0.2, rely=0.4, anchor=CENTER)

    msg3 = tk.Label(register_win, text='Authorization')
    msg3.place(relx=0.2, rely=0.6, anchor=CENTER)

    user_name = Entry(register_win, textvariable=register_user)
    user_name.place(relx=0.6, rely=0.2, anchor=CENTER, width=300)

    user_pass = Entry(register_win, textvariable=register_pass)
    user_pass.place(relx=0.6, rely=0.4, anchor=CENTER, width=300,)

    auth = Entry(register_win, textvariable=register_auth, )
    auth.place(relx=0.6, rely=0.6, anchor=CENTER, width=300)

    button = tk.Button(register_win, text='Sign Up', width=20, height=2, command=self.sign_up_backend,
                       activebackground="dark grey", activeforeground="red")
    button.place(relx=0.5, rely=0.8, anchor=CENTER)

    stop = tk.Button(register_win, text='EXIT', width=20, command=register_win.destroy,
                     bg="red", activebackground="red", )
    stop.place(relx=0.3, rely=1, anchor=SE)

  #----Function To Save the add user data in 2nd excel sheet---
  def sign_up_backend(self):
    auth = register_auth.get()
    if auth == "12345":
      current_row = sheet2.max_row
      current_col = sheet2.max_column
      sheet2.cell(row=current_row+1, column=1).value = register_user.get()
      sheet2.cell(row=current_row+1, column=2).value = register_pass.get()
      file.save(path.join(path.dirname(__file__), "data","student.xlsx"))
      messagebox.showinfo("Success", "User Successfuly Registered")

    else:
      messagebox.showerror("error", "Wrong Auth")

  #----Checking Input data in 2nd sheet----
  def sign_in_backend(self):
    a = login_user.get()  # getting string from login_user
    b = login_pass.get()  # getting string from login_pass
    last_row = sheet2.max_row+1
    if len(a) or len(b) != 0:
      for x in range(2, last_row):
        value1 = str(sheet2.cell(row=x, column=1).value)
        if a == value1:
          value2 = str(sheet2.cell(row=x, column=2).value)
          if b == value2:
            e = Management()

            break
          else:
            messagebox.showerror("error", "Wrong Password")
            break
      else:
        messagebox.showerror("error", "User not found Kindly register First")
    else:
      messagebox.showwarning("Warning", "All/Any fields are Enpty")

  #---This is the Class to Manage Dtudent Data---


class Management:
  #------Initializing the constructor of the class-----
  def __init__(self):
    win = Toplevel()
    win.geometry('750x600')
    win.configure(bg="#eaeaea")

    msg = Label(win, text="ABC School", font=75,
                width=100, height=4, )
    msg.place(relx=0.5, rely=0.135, anchor=S)

    msg1 = Label(win, text="Name", width=20, )
    msg1.place(relx=0.3, rely=0.2, anchor=CENTER)

    msg2 = Label(win, text="Father's Name", width=20, )
    msg2.place(relx=0.3, rely=0.25, anchor=CENTER)

    msg3 = Label(win, text="Enrollment no.", width=20, )
    msg3.place(relx=0.3, rely=0.3, anchor=CENTER)

    msg4 = Label(win, text="Class", width=20, )
    msg4.place(relx=0.3, rely=0.35, anchor=CENTER)

    msg5 = Label(win, text="Phone No.", width=20, )
    msg5.place(relx=0.3, rely=0.4, anchor=CENTER)

    msg6 = Label(win, text="Email", width=20, )
    msg6.place(relx=0.3, rely=0.45, anchor=CENTER)

    msg7 = Label(win, text="Address", width=20, )
    msg7.place(relx=0.3, rely=0.5, anchor=CENTER)

    ent1 = Entry(win, textvariable=ent1_var, width=50, )
    ent1.place(relx=0.65, rely=0.2, anchor=CENTER)

    ent2 = Entry(win, textvariable=ent2_var, width=50, )
    ent2.place(relx=0.65, rely=0.25, anchor=CENTER)

    ent3 = Entry(win, textvariable=ent3_var, width=50, )
    ent3.place(relx=0.65, rely=0.3, anchor=CENTER)

    ent4 = Entry(win, textvariable=ent4_var, width=50, )
    ent4.place(relx=0.65, rely=0.35, anchor=CENTER)

    ent5 = Entry(win, textvariable=ent5_var, width=50, )
    ent5.place(relx=0.65, rely=0.4, anchor=CENTER)

    ent6 = Entry(win, textvariable=ent6_var, width=50, )
    ent6.place(relx=0.65, rely=0.45, anchor=CENTER)

    ent7 = Entry(win, textvariable=ent7_var, width=50, )
    ent7.place(relx=0.65, rely=0.5, anchor=CENTER)

    check = tk.Button(win, text='Find', width=20, command=self.check_student,
                      bg="light grey", activebackground="green", )
    check.place(relx=0.3, rely=0.65, anchor=CENTER)

    add = tk.Button(win, text='Add', width=20, command=self.add_student,
                    bg="light grey", activebackground="green", )
    add.place(relx=0.7, rely=0.65, anchor=CENTER)

    delete = tk.Button(win, text='Delete', width=20, command=self.delete_student,
                       bg="light grey", activebackground="green", )
    delete.place(relx=0.5, rely=0.75, anchor=CENTER)

    stop = tk.Button(win, text='stop', width=75, command=win.destroy,
                     bg="red", activebackground="red", )
    stop.place(relx=0.5, rely=0.96, anchor=CENTER)

    #---Now Binding The Keys to Also Operate it Using Keyboard----
    def down1(Condition):
      ent2.focus_set()

    def down2(Condition):
      ent3.focus_set()

    def down3(Condition):
      ent4.focus_set()

    def down4(Condition):
      ent5.focus_set()

    def down5(Condition):
      ent6.focus_set()

    def down6(Condition):
      ent7.focus_set()

    def down7(Condition):
      check.focus_set()

    def down8(Condition):
      add.focus_set()

    def down9(Condition):
      delete.focus_set()

    def down10(Condition):
      stop.focus_set()

    def down11(Condition):
      ent1.focus_set()

    def ret1(Condition):
      self.check_student()

    def ret2(Condition):
      self.add_student()

    def ret3(Condition):
      self.delete_student()

    def ret4(Condition):
      win.destroy()

    ent1.bind("<Down>", down1)
    ent2.bind("<Down>", down2)
    ent3.bind("<Down>", down3)
    ent4.bind("<Down>", down4)
    ent5.bind("<Down>", down5)
    ent6.bind("<Down>", down6)
    ent7.bind("<Down>", down7)
    check.bind("<Down>", down8)
    add.bind("<Down>", down9)
    delete.bind("<Down>", down10)
    stop.bind("<Down>", down11)
    check.bind("<Return>", ret1)
    add.bind("<Return>", ret2)
    delete.bind("<Return>", ret3)
    stop.bind("<Return>", ret4)
  #====================================================

  #------Function To Find the Stored Values in the Sheet---
  def show(self, x):
    show = Toplevel()
    show.geometry('750x600')
    show.configure(bg="#eaeaea")

    msg = Label(show, text="The Student", font=75,
                width=100, height=4, relief=GROOVE)
    msg.place(relx=0.5, rely=0.135, anchor=S)

    msg1 = Label(show, text="Name", width=20, relief=GROOVE)
    msg1.place(relx=0.3, rely=0.2, anchor=CENTER)

    msg2 = Label(show, text="Father's Name", width=20, relief=GROOVE)
    msg2.place(relx=0.3, rely=0.25, anchor=CENTER)

    msg3 = Label(show, text="Enrollment no.", width=20, relief=GROOVE)
    msg3.place(relx=0.3, rely=0.3, anchor=CENTER)

    msg4 = Label(show, text="Class", width=20, relief=GROOVE)
    msg4.place(relx=0.3, rely=0.35, anchor=CENTER)

    msg5 = Label(show, text="Phone No.", width=20, relief=GROOVE)
    msg5.place(relx=0.3, rely=0.4, anchor=CENTER)

    msg6 = Label(show, text="Email", width=20, relief=GROOVE)
    msg6.place(relx=0.3, rely=0.45, anchor=CENTER)

    msg7 = Label(show, text="Address", width=20, relief=GROOVE)
    msg7.place(relx=0.3, rely=0.5, anchor=CENTER)

    ent1 = Entry(show, textvariable=ent1_var, width=50, relief=GROOVE)
    ent1.place(relx=0.65, rely=0.2, anchor=CENTER)

    ent2 = Entry(show, textvariable=ent2_var, width=50, relief=GROOVE)
    ent2.place(relx=0.65, rely=0.25, anchor=CENTER)

    ent3 = Entry(show, textvariable=ent3_var, width=50, relief=GROOVE)
    ent3.place(relx=0.65, rely=0.3, anchor=CENTER)

    ent4 = Entry(show, textvariable=ent4_var, width=50, relief=GROOVE)
    ent4.place(relx=0.65, rely=0.35, anchor=CENTER)

    ent5 = Entry(show, textvariable=ent5_var, width=50, relief=GROOVE)
    ent5.place(relx=0.65, rely=0.4, anchor=CENTER)

    ent6 = Entry(show, textvariable=ent6_var, width=50, relief=GROOVE)
    ent6.place(relx=0.65, rely=0.45, anchor=CENTER)

    ent7 = Entry(show, textvariable=ent7_var, width=50, relief=GROOVE)
    ent7.place(relx=0.65, rely=0.5, anchor=CENTER)
    #---------------------------------------------
    ent1_var.set(sheet1.cell(row=x, column=1).value)
    ent2_var.set(sheet1.cell(row=x, column=2).value)
    ent3_var.set(sheet1.cell(row=x, column=3).value)
    ent4_var.set(sheet1.cell(row=x, column=4).value)
    ent5_var.set(sheet1.cell(row=x, column=5).value)
    ent6_var.set(sheet1.cell(row=x, column=6).value)
    ent7_var.set(sheet1.cell(row=x, column=7).value)

    stop = tk.Button(show, text='stop', width=75, command=show.destroy,
                     bg="red", activebackground="red", relief=GROOVE)
    stop.place(relx=0.5, rely=0.96, anchor=CENTER)
    show.mainloop()

  #----Function To Store the input Values in the Sheet------
  def add_student(self):
    current_row = sheet1.max_row
    current_col = sheet1.max_column

    sheet1.cell(row=current_row+1, column=1).value = ent1_var.get()
    sheet1.cell(row=current_row+1, column=2).value = ent2_var.get()
    sheet1.cell(row=current_row+1, column=3).value = ent3_var.get()
    sheet1.cell(row=current_row+1, column=4).value = ent4_var.get()
    sheet1.cell(row=current_row+1, column=5).value = ent5_var.get()
    sheet1.cell(row=current_row+1, column=6).value = ent6_var.get()
    sheet1.cell(row=current_row+1, column=7).value = ent7_var.get()

    file.save(path.join(path.dirname(__file__), "data","student.xlsx"))
    messagebox.showinfo("Success", "Student is added \n To the Excle File")

  #----Function/TopLayer If Sudent is not found-------
  def student_not_found(self):
    student_not_found = Toplevel()
    student_not_found.geometry("350x250")
    student_not_found.title("User not Found ")
    student_not_found.configure(bg="yellow")
    stop = tk.Button(student_not_found, text='Student not Found ', width=30, height=2,
                     command=student_not_found.destroy, bg="red", activebackground="red", )
    stop.place(relx=0.5, rely=0.5, anchor=CENTER)

  #---Using Loop To Find The Input Data in sheet---
  def check_student(self):
    last = sheet1.max_row+1

    if len(ent3_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=3).value)
        if (str(ent3_var.get())) == value1:
          self.show(x)
          break
      else:
          self.student_not_found()

    elif len(ent5_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=5).value)
        if (str(ent5_var.get())) == value1:
          self.show(x)
          break
      else:
          self.student_not_found()

    elif len(ent1_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=1).value)
        if (str(ent1_var.get())) == value1:
          self.show(x)
          break
      else:
          self.student_not_found()

    elif len(ent2_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=2).value)
        if (str(ent2_var.get())) == value1:
          self.show(x)
          break
      else:
          self.student_not_found()

    elif len(ent4_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=4).value)
        if (str(ent4_var.get())) == value1:
          self.show(x)
          break
      else:
          self.student_not_found()

    elif len(ent6_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=6).value)
        if (str(ent6_var.get())) == value1:
          self.show(x)
          break
      else:
          self.student_not_found()

    elif len(ent7_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=7).value)
        if (str(ent_var.get())) == value1:
          self.show(x)
          break
      else:
          self.student_not_found()

    else:
      empty_field = Toplevel()
      empty_field.geometry("350x250")
      empty_field.title("Error")
      empty_field.configure(bg="yellow")
      stop = tk.Button(empty_field, text='Plese Provide atleast one field', width=30, height=2,
                       command=empty_field.destroy, bg="red", activebackground="red", )
      stop.place(relx=0.5, rely=0.5, anchor=CENTER)

  #-----Function to Empty The Cells in sheet----
  #-----And Restting the Variables------
  def remove_student(self, x):
    show = Toplevel()
    show.geometry('750x600')
    show.configure(bg="#eaeaea")

    msg = Label(show, text="The Student", font=75,
                width=100, height=4, relief=GROOVE)
    msg.place(relx=0.5, rely=0.135, anchor=S)

    msg1 = Label(show, text="Name", width=20, relief=GROOVE)
    msg1.place(relx=0.3, rely=0.2, anchor=CENTER)

    msg2 = Label(show, text="Father's Name", width=20, relief=GROOVE)
    msg2.place(relx=0.3, rely=0.25, anchor=CENTER)

    msg3 = Label(show, text="Enrollment no.", width=20, relief=GROOVE)
    msg3.place(relx=0.3, rely=0.3, anchor=CENTER)

    msg4 = Label(show, text="Class", width=20, relief=GROOVE)
    msg4.place(relx=0.3, rely=0.35, anchor=CENTER)

    msg5 = Label(show, text="Phone No.", width=20, relief=GROOVE)
    msg5.place(relx=0.3, rely=0.4, anchor=CENTER)

    msg6 = Label(show, text="Email", width=20, relief=GROOVE)
    msg6.place(relx=0.3, rely=0.45, anchor=CENTER)

    msg7 = Label(show, text="Address", width=20, relief=GROOVE)
    msg7.place(relx=0.3, rely=0.5, anchor=CENTER)

    ent1 = Entry(show, textvariable=ent1_var, width=50, relief=GROOVE)
    ent1.place(relx=0.65, rely=0.2, anchor=CENTER)

    ent2 = Entry(show, textvariable=ent2_var, width=50, relief=GROOVE)
    ent2.place(relx=0.65, rely=0.25, anchor=CENTER)

    ent3 = Entry(show, textvariable=ent3_var, width=50, relief=GROOVE)
    ent3.place(relx=0.65, rely=0.3, anchor=CENTER)

    ent4 = Entry(show, textvariable=ent4_var, width=50, relief=GROOVE)
    ent4.place(relx=0.65, rely=0.35, anchor=CENTER)

    ent5 = Entry(show, textvariable=ent5_var, width=50, relief=GROOVE)
    ent5.place(relx=0.65, rely=0.4, anchor=CENTER)

    ent6 = Entry(show, textvariable=ent6_var, width=50, relief=GROOVE)
    ent6.place(relx=0.65, rely=0.45, anchor=CENTER)

    ent7 = Entry(show, textvariable=ent7_var, width=50, relief=GROOVE)
    ent7.place(relx=0.65, rely=0.5, anchor=CENTER)

    #=================================================
    ent1_var.set(sheet.cell(row=x, column=1).value)
    ent2_var.set(sheet.cell(row=x, column=2).value)
    ent3_var.set(sheet.cell(row=x, column=3).value)
    ent4_var.set(sheet.cell(row=x, column=4).value)
    ent5_var.set(sheet.cell(row=x, column=5).value)
    ent6_var.set(sheet.cell(row=x, column=6).value)
    ent7_var.set(sheet.cell(row=x, column=7).value)

    def delete_student_data():
      (sheet.cell(row=x, column=1).value) = ""
      (sheet.cell(row=x, column=2).value) = ""
      (sheet.cell(row=x, column=3).value) = ""
      (sheet.cell(row=x, column=4).value) = ""
      (sheet.cell(row=x, column=5).value) = ""
      (sheet.cell(row=x, column=6).value) = ""
      (sheet.cell(row=x, column=7).value) = ""

      ent1_var.set("")
      ent2_var.set("")
      ent3_var.set("")
      ent4_var.set("")
      ent5_var.set("")
      ent6_var.set("")
      ent7_var.set("")

      file.save(path.join(path.dirname(__file__), "data","student.xlsx"))
      success = Toplevel()
      success.geometry("350x250")
      success.configure(bg="light green")
      success.title("Successfull Student deleted")
      sop = tk.Button(success, text='Success', width=25, height=2, command=success.destroy,
                      bg="green", activebackground="light grey", relief=GROOVE)
      sop.place(relx=0.5, rely=0.5, anchor=CENTER)

    stop = Button(show, text='delete', width=75, command=delete_student_data,
                  bg="red", activebackground="red", relief=GROOVE)
    stop.place(relx=0.5, rely=0.85, anchor=CENTER)

    stop = tk.Button(show, text='stop', width=75, command=show.destroy,
                     bg="red", activebackground="red", relief=GROOVE)
    stop.place(relx=0.5, rely=0.96, anchor=CENTER)

  #---Using Loop To Find The Input Data in sheet---
  def delete_student(self):
    last = sheet1.max_row+1

    if len(ent3_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=3).value)
        if (str(ent3_var.get())) == value1:
          self.remove_student(x)
          break
      else:
          self.student_not_found()

    elif len(ent5_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=5).value)
        if (str(ent5_var.get())) == value1:
          self.remove_student(x)
          break
      else:
          self.student_not_found()

    elif len(ent1_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=1).value)
        if (str(ent1_var.get())) == value1:
          self.remove_student(x)
          break
      else:
          self.student_not_found()

    elif len(ent2_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=2).value)
        if (str(ent2_var.get())) == value1:
          self.remove_student(x)
          break
      else:
          self.student_not_found()

    elif len(ent4_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=4).value)
        if (str(ent4_var.get())) == value1:
          self.remove_student(x)
          break
      else:
          self.student_not_found()

    elif len(ent6_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=6).value)
        if (str(ent6_var.get())) == value1:
          self.remove_student(x)
          break
      else:
          self.student_not_found()

    elif len(ent7_var.get()) != 0:
      for x in range(1, last):
        value1 = str(sheet1.cell(row=x, column=7).value)
        if (str(ent_var.get())) == value1:
          self.remove_student(x)
          break
      else:
          self.student_not_found()

    else:
      empty_field = Toplevel()
      empty_field.geometry("350x250")
      empty_field.title("Error")
      empty_field.configure(bg="yellow")
      stop = tk.Button(empty_field, text='Plese Provide atleast one field', width=30, height=2,
                       command=empty_field.destroy, bg="red", activebackground="red", )
      stop.place(relx=0.5, rely=0.5, anchor=CENTER)


#==================Starting Page======================
line = First_line()
line.page()

log = Login_sys()


regis = Menu(menu)
menu.add_cascade(label='options', menu=regis)
regis.add_command(label='Register', command=log.sign_up_frontend)

pic = Image.open(path.join(path.dirname(__file__), "data","login.jpg"))
picture = ImageTk.PhotoImage(pic)

image = tk.Button(app, width=200, height=200,
                  image=picture)
image.place(relx=0.5, rely=0.35, anchor=CENTER)

login = tk.Button(app, text='Sign In', width=35, height=3, command=log.sign_in_frontend,
                  activebackground="dark grey", activeforeground="red")
login.place(relx=0.5, rely=0.8, anchor=CENTER)

app.mainloop()
